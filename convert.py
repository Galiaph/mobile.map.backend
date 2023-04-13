#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Скрипт для заполнения базы данных из xlsx файла
"""

__author__ = "Vadim Shemyatin"

from openpyxl import load_workbook
import pymysql.cursors

def getBool(el):
    if el is None or el == 'Нет' or el == 'нет':
        return False
    return True

def getStatus(el):
    if el is None or el == 'Не в эфире':
        return False
    return True

if __name__ == "__main__":
    con = pymysql.connect(host='localhost', user='mobile', password='password', database='mobile', charset='utf8mb4', cursorclass=pymysql.cursors.DictCursor)
    with con:
        with con.cursor() as cursor:
           wb = load_workbook('./test7k.xlsx')
           #sheet = wb['МИР']
           sheet = wb['ХО']
           e = sheet.iter_rows()
           cells = list(e)
           for i in cells:
               sql = "INSERT base_station(bs_name, bs_latitude, bs_longitude, bs_comment, bs_operator, bs_2g, bs_3g, bs_4g, bs_status) VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s);"
               cursor.execute(sql, ('UNIT{0:03d}'.format(i[0].value), i[4].value, i[5].value, i[3].value, '6', 1, 0, 1, 1))
               #print('UNIT{0:03d}'.format(i[0].value), i[4].value, i[5].value, i[3].value)
        con.commit()

